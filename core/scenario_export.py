"""
core/scenario_export.py
=======================
Xuất kết quả chạy kịch bản grid (danh sách StepResult) ra:
  - CSV  (.csv)  — luôn xuất được, nguồn dữ liệu gốc, mở bằng Excel/LibreOffice.
  - Excel (.xlsx) — có header thông tin + tô màu trạng thái OK/LỖI (cần openpyxl).

Khác với core/report.py (dành cho MeasurementRecord của luồng cal tần số),
module này xử lý StepResult — bản ghi tổng quát theo (bước, thiết bị, hành động).

Không phụ thuộc Qt -> test được bằng pytest.
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from core.scenario_runner import StepResult, format_number_vi

log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    from openpyxl.utils import get_column_letter
    _EXCEL_OK = True
except ImportError:  # pragma: no cover
    _EXCEL_OK = False
    log.warning("openpyxl chưa cài — xuất .xlsx bị tắt, chỉ còn .csv")


# Thứ tự cột xuất ra.
FIELDS = ["step", "iteration", "timestamp", "device", "action",
          "value", "unit", "text", "status", "error"]
HEADERS_VI = ["Bước", "Lần lặp", "Thời gian", "Thiết bị", "Hành động",
              "Giá trị", "Đơn vị", "Chi tiết", "Trạng thái", "Lỗi"]


def result_to_row(r: StepResult) -> dict:
    """Quy đổi 1 StepResult thành dict phẳng để ghi CSV/Excel."""
    return {
        "step": r.step_index,
        "iteration": r.iteration or "",
        "timestamp": datetime.fromtimestamp(r.timestamp).strftime("%Y-%m-%d %H:%M:%S"),
        "device": r.device_key,
        "action": r.action,
        # Hiển thị kiểu VN (chấm nghìn, phẩy thập phân), không ký pháp khoa học.
        # Giá trị thô (raw response) vẫn còn nguyên ở cột "text" cho lệnh raw_scpi.
        "value": "" if r.value is None else format_number_vi(r.value),
        "unit": r.unit,
        "text": r.text,
        "status": "OK" if r.ok else "LỖI",
        "error": r.error,
    }


def export_csv(results: list[StepResult], path: str | Path) -> Path:
    """Ghi kết quả ra CSV (UTF-8 BOM để Excel hiển thị tiếng Việt đúng)."""
    path = Path(path)
    rows = [result_to_row(r) for r in results]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    log.info("CSV: %d dòng -> %s", len(rows), path)
    return path


def export_xlsx(
    results: list[StepResult],
    path: str | Path,
    meta: Optional[dict] = None,
) -> Path:
    """
    Ghi kết quả ra Excel có header thông tin + tô màu trạng thái.

    meta : dict tùy chọn, ví dụ {"scenario": "...", "operator": "...",
           "mode": "MOCK", "run_time": "2026-06-07 10:00"}.
    """
    if not _EXCEL_OK:
        raise RuntimeError("openpyxl chưa cài — không xuất được .xlsx (dùng CSV).")

    path = Path(path)
    meta = meta or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kết quả kịch bản"

    # ---- Khối thông tin ----
    info = [
        ("Kịch bản:", meta.get("scenario", "")),
        ("Người thực hiện:", meta.get("operator", "")),
        ("Chế độ:", meta.get("mode", "")),
        ("Thời gian chạy:", meta.get("run_time", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))),
        ("Tổng số kết quả:", len(results)),
        ("Số lỗi:", sum(1 for r in results if not r.ok)),
    ]
    for i, (label, value) in enumerate(info, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=value).font = Font(size=10)

    header_row = len(info) + 2

    # ---- Header bảng ----
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [6, 8, 20, 12, 18, 18, 8, 30, 12, 30]

    for c, (name, w) in enumerate(zip(HEADERS_VI, widths), start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Dữ liệu ----
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    err_fill = PatternFill("solid", fgColor="FFC7CE")
    for i, r in enumerate(results):
        row = header_row + 1 + i
        data = result_to_row(r)
        for c, key in enumerate(FIELDS, start=1):
            cell = ws.cell(row=row, column=c, value=data[key])
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if key == "status":
                cell.fill = ok_fill if r.ok else err_fill
                cell.font = Font(bold=True, size=10)
                cell.alignment = center

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    wb.save(path)
    log.info("XLSX: %d dòng -> %s", len(results), path)
    return path


def export(
    results: list[StepResult],
    path: str | Path,
    meta: Optional[dict] = None,
) -> Path:
    """Tự chọn định dạng theo đuôi file (.csv / .xlsx)."""
    if not results:
        raise ValueError("Không có kết quả để xuất.")
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return export_csv(results, path)
    if suffix in (".xlsx", ".xls"):
        return export_xlsx(results, path, meta)
    raise ValueError(f"Định dạng không hỗ trợ: {suffix} (dùng .csv hoặc .xlsx)")
