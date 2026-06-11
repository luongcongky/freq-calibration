"""
unit_test/test_scenario_export.py
=================================
Test xuất kết quả kịch bản (core/scenario_export.py) ra CSV/XLSX.
"""

import csv

import pytest

from core.scenario import Scenario, ScenarioStep
from core.scenario_runner import ScenarioRunner, StepResult
from core import scenario_export as sx


def _sample_results() -> list[StepResult]:
    scn = Scenario(nodes=[
        ScenarioStep(action="identify", devices=["CNT91", "N1913A"]),
        ScenarioStep(action="measure_frequency", devices=["CNT91"]),
        ScenarioStep(action="measure_power", devices=["N1913A"]),
    ])
    return ScenarioRunner(mock=True, settle_wait=False).run(scn)


def test_export_csv(tmp_path):
    results = _sample_results()
    path = tmp_path / "out.csv"
    sx.export_csv(results, path)
    assert path.exists()

    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == len(results)
    assert rows[0]["action"] == "identify"
    assert set(sx.FIELDS).issubset(rows[0].keys())
    # dòng đo tần số có value (định dạng VN: chấm nghìn, phẩy thập phân) + unit Hz
    freq = [r for r in rows if r["action"] == "measure_frequency"][0]
    assert freq["unit"] == "Hz"
    raw = float(freq["value"].replace(".", "").replace(",", "."))   # gỡ định dạng VN -> số
    assert raw > 0
    assert "e" not in freq["value"].lower()                         # không còn ký pháp khoa học


def test_export_dispatch_by_extension(tmp_path):
    results = _sample_results()
    p = sx.export(results, tmp_path / "x.csv")
    assert p.suffix == ".csv" and p.exists()


def test_export_empty_raises(tmp_path):
    with pytest.raises(ValueError, match="Không có kết quả"):
        sx.export([], tmp_path / "x.csv")


def test_export_bad_extension(tmp_path):
    with pytest.raises(ValueError, match="không hỗ trợ"):
        sx.export(_sample_results(), tmp_path / "x.pdf")


@pytest.mark.skipif(not sx._EXCEL_OK, reason="openpyxl chưa cài")
def test_export_xlsx(tmp_path):
    import openpyxl
    results = _sample_results()
    path = tmp_path / "out.xlsx"
    sx.export_xlsx(results, path, meta={"scenario": "Demo", "mode": "MOCK", "operator": "KTV"})
    assert path.exists()

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # header thông tin ở vài dòng đầu + bảng bên dưới; kiểm tra có chữ 'Kịch bản' và 'Bước'
    all_vals = [ws.cell(row=r, column=c).value
                for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1)]
    assert "Kịch bản:" in all_vals     # nhãn khối thông tin
    assert "Bước" in all_vals          # header bảng
    assert "OK" in all_vals            # cột trạng thái
